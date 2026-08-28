"""
export.py
Builds an Excel workbook of a user's reading so they can keep their data.

Sheets:
  Birth chart   tropical and sidereal placements plus angles
  Dasha         the Vimshottari mahadasha timeline
  Today         the day's factors
  Guidance      today's verdict for every life area
"""

from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

import astro
import domains

HEADER_FILL = PatternFill("solid", fgColor="2F4B7C")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=13)


def _header(ws, row, cols):
    for i, c in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left")


def _autofit(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w


def build_workbook(natal, name, transit_trop, jd, weekday_index):
    wb = openpyxl.Workbook()

    # --- Birth chart ---
    ws = wb.active
    ws.title = "Birth chart"
    ws["A1"] = f"Sama reading for {name or 'you'}"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Tropical (Western) and sidereal (Vedic) placements"

    ws["A4"] = "Ascendant (tropical)"
    ws["B4"] = f"{astro.SIGNS[int(natal['asc_tropical']//30)]} " \
               f"{round(natal['asc_tropical']%30,2)} deg"
    ws["A5"] = "Midheaven (tropical)"
    ws["B5"] = f"{astro.SIGNS[int(natal['mc_tropical']//30)]} " \
               f"{round(natal['mc_tropical']%30,2)} deg"
    ws["A6"] = "Ascendant (sidereal)"
    ws["B6"] = f"{natal['asc_sidereal_sign']} " \
               f"{round(natal['asc_sidereal']%30,2)} deg"
    ws["A7"] = "Moon nakshatra"
    ws["B7"] = f"{natal['moon_nakshatra']['name']} " \
               f"(lord {natal['moon_nakshatra']['lord']})"

    _header(ws, 9, ["Planet", "Tropical sign", "Deg", "House",
                    "Sidereal sign", "Retro"])
    r = 10
    order = ["Sun", "Moon", "Mercury", "Venus", "Mars", "Jupiter",
             "Saturn", "Uranus", "Neptune", "Pluto", "Rahu", "Ketu"]
    for p in order:
        tp = natal["tropical"][p]
        sp = natal["sidereal"][p]
        house = astro.house_of(tp["lon"], natal["cusps"])
        ws.cell(r, 1, p)
        ws.cell(r, 2, tp["sign"])
        ws.cell(r, 3, round(tp["deg_in_sign"], 2))
        ws.cell(r, 4, house)
        ws.cell(r, 5, sp["sign"])
        ws.cell(r, 6, "yes" if tp["retro"] else "")
        r += 1
    _autofit(ws, [16, 14, 8, 7, 14, 8])

    # --- Dasha ---
    ws2 = wb.create_sheet("Dasha")
    ws2["A1"] = "Vimshottari Dasha timeline"
    ws2["A1"].font = TITLE_FONT
    _header(ws2, 3, ["Mahadasha lord", "Start", "End", "Years"])
    r = 4
    for period in natal["dasha_periods"]:
        ws2.cell(r, 1, period["lord"])
        ws2.cell(r, 2, period["start"].date().isoformat())
        ws2.cell(r, 3, period["end"].date().isoformat())
        ws2.cell(r, 4, period["years"])
        r += 1
    _autofit(ws2, [18, 14, 14, 8])

    # --- Today ---
    ws3 = wb.create_sheet("Today")
    ws3["A1"] = "Today's sky"
    ws3["A1"].font = TITLE_FONT
    pan = astro.panchanga(jd, weekday_index)
    rows = [
        ("Weekday", pan["vara"]),
        ("Tithi", f"{pan['tithi']} ({pan['paksha']})"),
        ("Nakshatra", f"{pan['nakshatra']} pada {pan['nakshatra_pada']}"),
        ("Yoga", pan["yoga"]),
        ("Moon sidereal sign", pan["moon_sidereal_sign"]),
    ]
    r = 3
    for k, v in rows:
        ws3.cell(r, 1, k).font = Font(bold=True)
        ws3.cell(r, 2, v)
        r += 1
    _header(ws3, r + 1, ["Transiting", "Aspect", "Natal", "Orb"])
    r += 2
    hits = astro.transit_aspects(transit_trop, natal["tropical"], orb=5.0)
    for h in hits[:15]:
        ws3.cell(r, 1, h["transiting"])
        ws3.cell(r, 2, h["aspect"])
        ws3.cell(r, 3, h["natal"])
        ws3.cell(r, 4, h["orb"])
        r += 1
    _autofit(ws3, [16, 14, 12, 8])

    # --- Guidance ---
    ws4 = wb.create_sheet("Guidance")
    ws4["A1"] = "Today's life-area guidance"
    ws4["A1"].font = TITLE_FONT
    ws4["A2"] = ("Reflective timing only. Not medical, financial, or legal "
                 "advice.")
    _header(ws4, 4, ["Area", "Verdict", "Score"])
    r = 5
    for row in domains.scan_all(natal, transit_trop, jd, weekday_index,
                                natal["cusps"]):
        ws4.cell(r, 1, row["domain"])
        ws4.cell(r, 2, row["verdict"])
        ws4.cell(r, 3, row["score"])
        r += 1
    _autofit(ws4, [26, 28, 8])

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
